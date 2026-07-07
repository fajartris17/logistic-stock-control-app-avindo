from flask import Flask, render_template, jsonify, request, send_file, send_from_directory
from pathlib import Path
from datetime import datetime
from uuid import uuid4
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from werkzeug.utils import secure_filename
import re

BASE_DIR = Path(__file__).resolve().parent
DATABASE_DIR = BASE_DIR / 'database'
UPLOAD_DIR = BASE_DIR / 'uploads'
DB_PATH = DATABASE_DIR / 'logistic_stock_database.xlsx'

ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.webp', '.pdf'}
MAX_UPLOAD_MB = 10

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024

DAY_NAMES = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']

SHEET_HEADERS = {
    'DATABASE_BARANG_MASUK': [
        'ID Transaksi', 'Tanggal Input', 'Jam Input', 'Code Number', 'Nama Barang', 'Jenis Barang',
        'QTY Masuk', 'Satuan', 'Total QTY', 'Harga Satuan', 'Total Harga', 'Supplier / Asal Barang',
        'Nama File Foto Barang', 'Path Foto Barang', 'Nama File Nota', 'Path Foto Nota',
        'Status Koreksi', 'Keterangan', 'User Input'
    ],
    'DATABASE_BARANG_KELUAR': [
        'ID Transaksi', 'Tanggal Input', 'Jam Input', 'Code Number', 'Nama Barang', 'Jenis Barang',
        'QTY Keluar', 'Satuan', 'Total QTY', 'Harga Satuan', 'Total Harga', 'Tujuan / Pengguna Barang',
        'Nama File Foto Barang', 'Path Foto Barang', 'Nama File Bukti Keluar', 'Path Foto Bukti Keluar',
        'Status Koreksi', 'Keterangan', 'User Input'
    ],
    'DATABASE_STOK': [
        'Code Number', 'Nama Barang', 'Jenis Barang', 'Total QTY Masuk', 'Total QTY Keluar', 'Stok Akhir',
        'Satuan', 'Harga Satuan Terakhir', 'Estimasi Nilai Stok', 'Minimum Stok', 'Status Stok',
        'Foto Barang Terakhir', 'Keterangan'
    ],
    'MASTER_BARANG': [
        'Code Number', 'Nama Barang', 'Jenis Barang', 'Satuan', 'Harga Satuan Terakhir', 'Minimum Stok', 'Keterangan'
    ],
    'LOG_EDIT_DATA': [
        'ID Edit', 'Tanggal Edit', 'Jam Edit', 'ID Transaksi', 'Field yang Diedit', 'Data Lama', 'Data Baru',
        'User Edit', 'Keterangan Edit'
    ],
    'DATABASE_NOTA_OCR': [
        'ID OCR', 'Tanggal Upload', 'Jam Upload', 'Nama File Nota', 'Path Foto Nota', 'Tanggal Nota',
        'Nama Supplier / Toko', 'Nomor Nota', 'Nama Barang OCR', 'QTY OCR', 'Satuan OCR',
        'Harga Satuan OCR', 'Total Harga OCR', 'Total Nota OCR', 'Status Pembacaan', 'Keterangan OCR'
    ],
    'LOG_KOREKSI_DATA': [
        'ID Koreksi', 'Tanggal Koreksi', 'Jam Koreksi', 'ID Transaksi', 'Nama Barang Input', 'Nama Barang Nota',
        'QTY Input', 'QTY Nota', 'Harga Satuan Input', 'Harga Satuan Nota', 'Total Harga Input',
        'Total Harga Nota', 'Status Koreksi', 'Tindakan User', 'Keterangan'
    ]
}

ID_PREFIX_BY_SHEET = {
    'DATABASE_BARANG_MASUK': 'BM',
    'DATABASE_BARANG_KELUAR': 'BK',
    'LOG_EDIT_DATA': 'EDIT',
    'DATABASE_NOTA_OCR': 'OCR',
    'LOG_KOREKSI_DATA': 'KOREKSI'
}


def now_date():
    return datetime.now().strftime('%Y-%m-%d')


def now_time():
    return datetime.now().strftime('%H:%M:%S')


def today_key():
    return datetime.now().strftime('%Y%m%d')


def to_float(value):
    try:
        if value is None or value == '':
            return 0.0
        text = str(value).strip()
        if ',' in text:
            text = text.replace('.', '').replace(',', '.')
        return float(text)
    except Exception:
        return 0.0


def rupiah(value):
    return 'Rp {:,.0f}'.format(to_float(value)).replace(',', '.')


def form_or_json():
    if request.files or request.form:
        return request.form
    return request.get_json(silent=True) or {}


def normalize_code(value):
    return str(value or '').strip().upper()


def public_upload_path(path_obj):
    if not path_obj:
        return ''
    rel = path_obj.relative_to(UPLOAD_DIR).as_posix()
    return f'/uploads/{rel}'


def style_sheet(ws):
    header_fill = PatternFill('solid', fgColor='0B1F3A')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='D9E2EC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if ws.max_row < 1:
        return

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 34

    for idx, cell in enumerate(ws[1], start=1):
        header = str(cell.value or '')
        h = header.lower()
        width = 14
        if any(k in h for k in ['nama', 'keterangan', 'supplier', 'tujuan', 'path', 'field', 'foto', 'nota', 'bukti']):
            width = 26
        if 'id' in h or 'code' in h:
            width = 19
        if 'tanggal' in h or 'jam' in h:
            width = 16
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)


def add_or_update_table(ws):
    if ws.max_row < 1 or ws.max_column < 1:
        return
    table_name = re.sub(r'[^A-Za-z0-9_]', '', ws.title) + 'Table'
    ref = f'A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 2)}'
    try:
        if table_name in ws.tables:
            ws.tables[table_name].ref = ref
            return
        tab = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    except Exception:
        pass


def ensure_database():
    """Membuat/merapikan database Excel tanpa menghapus data lama."""
    DATABASE_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

    if DB_PATH.exists():
        wb = load_workbook(DB_PATH)
    else:
        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)

    for sheet_name, headers in SHEET_HEADERS.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
        else:
            ws = wb[sheet_name]
            existing_headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
            if not any(existing_headers):
                ws.append(headers)
                existing_headers = headers
            for header in headers:
                if header not in existing_headers:
                    ws.cell(row=1, column=ws.max_column + 1, value=header)
        style_sheet(wb[sheet_name])
        add_or_update_table(wb[sheet_name])

    wb.save(DB_PATH)


def read_sheet(sheet_name):
    ensure_database()
    wb = load_workbook(DB_PATH, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or '') for h in rows[0]]
    data = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        item['_row'] = row_idx
        data.append(item)
    return data


def append_dict_row(sheet_name, data):
    ensure_database()
    wb = load_workbook(DB_PATH)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    ws.append([data.get(h, '') for h in headers])
    style_sheet(ws)
    add_or_update_table(ws)
    wb.save(DB_PATH)


def next_id(sheet_name, prefix=None):
    prefix = prefix or ID_PREFIX_BY_SHEET.get(sheet_name, 'ID')
    date_key = today_key()
    pattern = re.compile(rf'^{re.escape(prefix)}-{date_key}-(\d{{4}})$')
    max_no = 0
    for row in read_sheet(sheet_name):
        value = str(row.get('ID Transaksi') or row.get('ID Edit') or row.get('ID OCR') or row.get('ID Koreksi') or '')
        match = pattern.match(value)
        if match:
            max_no = max(max_no, int(match.group(1)))
    return f'{prefix}-{date_key}-{max_no + 1:04d}'


def save_upload(field_name, subfolder):
    file = request.files.get(field_name)
    if not file or not file.filename:
        return '', ''

    original = secure_filename(file.filename)
    ext = Path(original).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(f'Format file {original} tidak diizinkan. Gunakan JPG, JPEG, PNG, WEBP, atau PDF.')

    folder = UPLOAD_DIR / subfolder / today_key()
    folder.mkdir(parents=True, exist_ok=True)
    safe_name = f'{today_key()}_{uuid4().hex[:8]}_{original}'
    target = folder / safe_name
    file.save(target)

    max_bytes = MAX_UPLOAD_MB * 1024 * 1024
    if target.stat().st_size > max_bytes:
        target.unlink(missing_ok=True)
        raise ValueError(f'Ukuran file {original} melebihi {MAX_UPLOAD_MB} MB.')

    return original, public_upload_path(target)


def upsert_master_barang(code, nama, jenis, satuan, harga, minimum=0, keterangan=''):
    if not code:
        return
    ensure_database()
    wb = load_workbook(DB_PATH)
    ws = wb['MASTER_BARANG']
    headers = [c.value for c in ws[1]]
    code_idx = headers.index('Code Number') + 1
    found_row = None

    for r in range(2, ws.max_row + 1):
        if normalize_code(ws.cell(r, code_idx).value) == code:
            found_row = r
            break

    row_data = {
        'Code Number': code,
        'Nama Barang': nama,
        'Jenis Barang': jenis,
        'Satuan': satuan,
        'Harga Satuan Terakhir': harga,
        'Minimum Stok': minimum,
        'Keterangan': keterangan
    }

    if found_row:
        for h, v in row_data.items():
            if v not in [None, '']:
                ws.cell(found_row, headers.index(h) + 1, value=v)
    else:
        ws.append([row_data.get(h, '') for h in headers])

    style_sheet(ws)
    add_or_update_table(ws)
    wb.save(DB_PATH)


def get_master_map():
    return {normalize_code(r.get('Code Number')): r for r in read_sheet('MASTER_BARANG')}


def rebuild_stock():
    ensure_database()
    masuk = read_sheet('DATABASE_BARANG_MASUK')
    keluar = read_sheet('DATABASE_BARANG_KELUAR')
    master = get_master_map()
    stock = {}

    def get_item(code, fallback=None):
        fallback = fallback or {}
        m = master.get(code, {})
        return stock.setdefault(code, {
            'Code Number': code,
            'Nama Barang': m.get('Nama Barang') or fallback.get('Nama Barang') or '',
            'Jenis Barang': m.get('Jenis Barang') or fallback.get('Jenis Barang') or '',
            'Total QTY Masuk': 0.0,
            'Total QTY Keluar': 0.0,
            'Stok Akhir': 0.0,
            'Satuan': m.get('Satuan') or fallback.get('Satuan') or '',
            'Harga Satuan Terakhir': to_float(m.get('Harga Satuan Terakhir') or fallback.get('Harga Satuan') or 0),
            'Estimasi Nilai Stok': 0.0,
            'Minimum Stok': to_float(m.get('Minimum Stok') or 0),
            'Status Stok': 'TERSEDIA',
            'Foto Barang Terakhir': '',
            'Keterangan': m.get('Keterangan') or ''
        })

    for row in masuk:
        code = normalize_code(row.get('Code Number'))
        if not code:
            continue
        item = get_item(code, row)
        qty = to_float(row.get('Total QTY') or row.get('QTY Masuk'))
        harga = to_float(row.get('Harga Satuan'))
        item['Total QTY Masuk'] += qty
        if harga:
            item['Harga Satuan Terakhir'] = harga
        item['Nama Barang'] = row.get('Nama Barang') or item['Nama Barang']
        item['Jenis Barang'] = row.get('Jenis Barang') or item['Jenis Barang']
        item['Satuan'] = row.get('Satuan') or item['Satuan']
        item['Foto Barang Terakhir'] = row.get('Path Foto Barang') or row.get('Nama File Foto Barang') or item['Foto Barang Terakhir']

    for row in keluar:
        code = normalize_code(row.get('Code Number'))
        if not code:
            continue
        item = get_item(code, row)
        qty = to_float(row.get('Total QTY') or row.get('QTY Keluar'))
        item['Total QTY Keluar'] += qty
        item['Nama Barang'] = row.get('Nama Barang') or item['Nama Barang']
        item['Jenis Barang'] = row.get('Jenis Barang') or item['Jenis Barang']
        item['Satuan'] = row.get('Satuan') or item['Satuan']

    for item in stock.values():
        item['Stok Akhir'] = item['Total QTY Masuk'] - item['Total QTY Keluar']
        item['Estimasi Nilai Stok'] = item['Stok Akhir'] * item['Harga Satuan Terakhir']
        if item['Stok Akhir'] <= 0:
            item['Status Stok'] = 'STOK HABIS'
        elif item['Minimum Stok'] and item['Stok Akhir'] <= item['Minimum Stok']:
            item['Status Stok'] = 'STOK MINIMUM'
        else:
            item['Status Stok'] = 'TERSEDIA'

    wb = load_workbook(DB_PATH)
    ws = wb['DATABASE_STOK']
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    headers = [c.value for c in ws[1]]
    for _, item in sorted(stock.items()):
        ws.append([item.get(h, '') for h in headers])
    style_sheet(ws)
    add_or_update_table(ws)
    wb.save(DB_PATH)
    return list(stock.values())


def current_stock_for(code):
    for item in rebuild_stock():
        if normalize_code(item.get('Code Number')) == code:
            return to_float(item.get('Stok Akhir'))
    return 0.0


def filter_rows(rows, start='', end='', q=''):
    q = str(q or '').strip().lower()
    out = []
    for row in rows:
        tanggal = str(row.get('Tanggal') or row.get('Tanggal Input') or '')
        text = ' '.join(str(v or '') for v in row.values()).lower()
        if start and tanggal < start:
            continue
        if end and tanggal > end:
            continue
        if q and q not in text:
            continue
        out.append(row)
    return out


def build_laporan():
    rows = []
    for r in read_sheet('DATABASE_BARANG_MASUK'):
        rows.append({
            'Tanggal': r.get('Tanggal Input'), 'Jam': r.get('Jam Input'), 'Jenis Transaksi': 'BARANG MASUK',
            'ID Transaksi': r.get('ID Transaksi'), 'Code Number': r.get('Code Number'), 'Nama Barang': r.get('Nama Barang'),
            'Jenis Barang': r.get('Jenis Barang'), 'Qty': r.get('Total QTY') or r.get('QTY Masuk'), 'Satuan': r.get('Satuan'),
            'Nilai': r.get('Total Harga'), 'Pihak': r.get('Supplier / Asal Barang'), 'Status Koreksi': r.get('Status Koreksi'),
            'Keterangan': r.get('Keterangan'), 'Foto': r.get('Path Foto Barang') or r.get('Nama File Foto Barang'),
            'Bukti': r.get('Path Foto Nota') or r.get('Nama File Nota')
        })
    for r in read_sheet('DATABASE_BARANG_KELUAR'):
        rows.append({
            'Tanggal': r.get('Tanggal Input'), 'Jam': r.get('Jam Input'), 'Jenis Transaksi': 'BARANG KELUAR',
            'ID Transaksi': r.get('ID Transaksi'), 'Code Number': r.get('Code Number'), 'Nama Barang': r.get('Nama Barang'),
            'Jenis Barang': r.get('Jenis Barang'), 'Qty': r.get('Total QTY') or r.get('QTY Keluar'), 'Satuan': r.get('Satuan'),
            'Nilai': r.get('Total Harga'), 'Pihak': r.get('Tujuan / Pengguna Barang'), 'Status Koreksi': r.get('Status Koreksi'),
            'Keterangan': r.get('Keterangan'), 'Foto': r.get('Path Foto Barang') or r.get('Nama File Foto Barang'),
            'Bukti': r.get('Path Foto Bukti Keluar') or r.get('Nama File Bukti Keluar')
        })
    rows.sort(key=lambda x: str(x.get('Tanggal') or '') + ' ' + str(x.get('Jam') or ''))
    return rows


def find_transaction(txid):
    for sheet in ['DATABASE_BARANG_MASUK', 'DATABASE_BARANG_KELUAR']:
        for row in read_sheet(sheet):
            if str(row.get('ID Transaksi') or '') == txid:
                return sheet, row
    return None, None


def update_transaction_fields(txid, updates, user='Admin', note='Edit manual dari aplikasi'):
    sheet, row = find_transaction(txid)
    if not row:
        return False, 'ID transaksi tidak ditemukan.'

    wb = load_workbook(DB_PATH)
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    row_num = int(row['_row'])
    editable = {
        'Nama Barang', 'Jenis Barang', 'Satuan', 'Harga Satuan', 'Supplier / Asal Barang',
        'Tujuan / Pengguna Barang', 'Status Koreksi', 'Keterangan', 'User Input'
    }

    # QTY alias.
    if sheet == 'DATABASE_BARANG_MASUK':
        editable.update({'QTY Masuk', 'Total QTY'})
    else:
        editable.update({'QTY Keluar', 'Total QTY'})

    changes = []
    for field, new_value in updates.items():
        if field not in editable or field not in headers:
            continue
        col = headers.index(field) + 1
        old_value = ws.cell(row_num, col).value
        if str(old_value or '') != str(new_value or ''):
            ws.cell(row_num, col, value=new_value)
            changes.append((field, old_value, new_value))

    # Recalculate total harga if qty or price changed.
    headers_after = [c.value for c in ws[1]]
    qty_header = 'QTY Masuk' if sheet == 'DATABASE_BARANG_MASUK' else 'QTY Keluar'
    qty = to_float(ws.cell(row_num, headers_after.index(qty_header) + 1).value)
    harga = to_float(ws.cell(row_num, headers_after.index('Harga Satuan') + 1).value)
    total_col = headers_after.index('Total Harga') + 1
    total_qty_col = headers_after.index('Total QTY') + 1
    old_total = ws.cell(row_num, total_col).value
    ws.cell(row_num, total_qty_col, value=qty)
    ws.cell(row_num, total_col, value=qty * harga)
    if str(old_total or '') != str(qty * harga):
        changes.append(('Total Harga', old_total, qty * harga))

    style_sheet(ws)
    add_or_update_table(ws)
    wb.save(DB_PATH)

    for field, old_value, new_value in changes:
        append_dict_row('LOG_EDIT_DATA', {
            'ID Edit': next_id('LOG_EDIT_DATA', 'EDIT'),
            'Tanggal Edit': now_date(),
            'Jam Edit': now_time(),
            'ID Transaksi': txid,
            'Field yang Diedit': field,
            'Data Lama': old_value,
            'Data Baru': new_value,
            'User Edit': user,
            'Keterangan Edit': note
        })

    rebuild_stock()
    return True, f'{len(changes)} field diperbarui.'


@app.route('/')
def index():
    ensure_database()
    return render_template('index.html')


@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)


@app.route('/api/dashboard')
def dashboard():
    ensure_database()
    rebuild_stock()
    today = datetime.now()
    masuk = read_sheet('DATABASE_BARANG_MASUK')
    keluar = read_sheet('DATABASE_BARANG_KELUAR')
    stok = read_sheet('DATABASE_STOK')
    nilai_masuk = sum(to_float(r.get('Total Harga')) for r in masuk)
    nilai_keluar = sum(to_float(r.get('Total Harga')) for r in keluar)
    minimum = [r for r in stok if r.get('Status Stok') == 'STOK MINIMUM']
    habis = [r for r in stok if r.get('Status Stok') == 'STOK HABIS']

    return jsonify({
        'tanggal': today.strftime('%d-%m-%Y'),
        'tanggal_iso': today.strftime('%Y-%m-%d'),
        'hari': DAY_NAMES[today.weekday()],
        'total_jenis_barang': len(stok),
        'total_barang_masuk': sum(to_float(r.get('Total QTY') or r.get('QTY Masuk')) for r in masuk),
        'total_barang_keluar': sum(to_float(r.get('Total QTY') or r.get('QTY Keluar')) for r in keluar),
        'total_stok_tersedia': sum(to_float(r.get('Stok Akhir')) for r in stok),
        'total_nilai_masuk': nilai_masuk,
        'total_nilai_keluar': nilai_keluar,
        'total_nilai_masuk_fmt': rupiah(nilai_masuk),
        'total_nilai_keluar_fmt': rupiah(nilai_keluar),
        'stok_minimum': len(minimum),
        'stok_habis': len(habis),
        'recent_stock': stok[-8:]
    })


@app.route('/api/stok')
def api_stok():
    q = request.args.get('q', '')
    data = rebuild_stock()
    if q:
        q_lower = q.lower()
        data = [r for r in data if q_lower in ' '.join(str(v or '') for v in r.values()).lower()]
    return jsonify(data)


@app.route('/api/barang-masuk', methods=['POST'])
def api_barang_masuk():
    try:
        payload = form_or_json()
        qty = to_float(payload.get('qty'))
        harga = to_float(payload.get('harga_satuan'))
        code = normalize_code(payload.get('kode_barang'))
        nama = str(payload.get('nama_barang') or '').strip()

        if not code or not nama or qty <= 0:
            return jsonify({'ok': False, 'message': 'Code Number, Nama Barang, dan QTY Masuk wajib diisi benar.'}), 400

        total = qty * harga
        foto_name, foto_path = save_upload('foto_barang', 'barang')
        nota_name, nota_path = save_upload('foto_nota', 'nota')

        row = {
            'ID Transaksi': next_id('DATABASE_BARANG_MASUK', 'BM'),
            'Tanggal Input': payload.get('tanggal') or now_date(),
            'Jam Input': now_time(),
            'Code Number': code,
            'Nama Barang': nama,
            'Jenis Barang': payload.get('kategori') or payload.get('jenis_barang') or '',
            'QTY Masuk': qty,
            'Satuan': payload.get('satuan') or '',
            'Total QTY': qty,
            'Harga Satuan': harga,
            'Total Harga': total,
            'Supplier / Asal Barang': payload.get('supplier') or '',
            'Nama File Foto Barang': foto_name,
            'Path Foto Barang': foto_path,
            'Nama File Nota': nota_name,
            'Path Foto Nota': nota_path,
            'Status Koreksi': 'BELUM DICEK',
            'Keterangan': payload.get('keterangan') or '',
            'User Input': payload.get('user_input') or 'Admin'
        }
        append_dict_row('DATABASE_BARANG_MASUK', row)
        upsert_master_barang(code, row['Nama Barang'], row['Jenis Barang'], row['Satuan'], harga,
                             to_float(payload.get('minimum_stok')), row['Keterangan'])
        rebuild_stock()
        return jsonify({'ok': True, 'message': f"Barang masuk tersimpan dengan ID {row['ID Transaksi']}", 'id': row['ID Transaksi'], 'total': total, 'total_fmt': rupiah(total)})
    except ValueError as exc:
        return jsonify({'ok': False, 'message': str(exc)}), 400
    except Exception as exc:
        return jsonify({'ok': False, 'message': f'Gagal menyimpan barang masuk: {exc}'}), 500


@app.route('/api/barang-keluar', methods=['POST'])
def api_barang_keluar():
    try:
        payload = form_or_json()
        qty = to_float(payload.get('qty'))
        harga = to_float(payload.get('harga_satuan'))
        code = normalize_code(payload.get('kode_barang'))
        nama = str(payload.get('nama_barang') or '').strip()

        if not code or not nama or qty <= 0:
            return jsonify({'ok': False, 'message': 'Code Number, Nama Barang, dan QTY Keluar wajib diisi benar.'}), 400

        stok_saat_ini = current_stock_for(code)
        if qty > stok_saat_ini:
            return jsonify({'ok': False, 'message': f'Stok tidak cukup. Stok saat ini {stok_saat_ini:g}, diminta keluar {qty:g}.'}), 400

        total = qty * harga
        foto_name, foto_path = save_upload('foto_barang', 'barang')
        bukti_name, bukti_path = save_upload('foto_bukti_keluar', 'bukti_keluar')

        row = {
            'ID Transaksi': next_id('DATABASE_BARANG_KELUAR', 'BK'),
            'Tanggal Input': payload.get('tanggal') or now_date(),
            'Jam Input': now_time(),
            'Code Number': code,
            'Nama Barang': nama,
            'Jenis Barang': payload.get('kategori') or payload.get('jenis_barang') or '',
            'QTY Keluar': qty,
            'Satuan': payload.get('satuan') or '',
            'Total QTY': qty,
            'Harga Satuan': harga,
            'Total Harga': total,
            'Tujuan / Pengguna Barang': payload.get('penerima') or payload.get('tujuan') or '',
            'Nama File Foto Barang': foto_name,
            'Path Foto Barang': foto_path,
            'Nama File Bukti Keluar': bukti_name,
            'Path Foto Bukti Keluar': bukti_path,
            'Status Koreksi': 'BELUM DICEK',
            'Keterangan': payload.get('keterangan') or '',
            'User Input': payload.get('user_input') or 'Admin'
        }
        append_dict_row('DATABASE_BARANG_KELUAR', row)
        upsert_master_barang(code, row['Nama Barang'], row['Jenis Barang'], row['Satuan'], harga,
                             to_float(payload.get('minimum_stok')), row['Keterangan'])
        rebuild_stock()
        return jsonify({'ok': True, 'message': f"Barang keluar tersimpan dengan ID {row['ID Transaksi']}", 'id': row['ID Transaksi'], 'total': total, 'total_fmt': rupiah(total)})
    except ValueError as exc:
        return jsonify({'ok': False, 'message': str(exc)}), 400
    except Exception as exc:
        return jsonify({'ok': False, 'message': f'Gagal menyimpan barang keluar: {exc}'}), 500


@app.route('/api/laporan')
def api_laporan():
    rebuild_stock()
    rows = build_laporan()
    rows = filter_rows(rows, request.args.get('start', ''), request.args.get('end', ''), request.args.get('q', ''))
    return jsonify(rows)


@app.route('/api/transaksi/<txid>')
def api_get_transaksi(txid):
    sheet, row = find_transaction(txid)
    if not row:
        return jsonify({'ok': False, 'message': 'ID transaksi tidak ditemukan.'}), 404
    row['_sheet'] = sheet
    return jsonify({'ok': True, 'data': row})


@app.route('/api/transaksi/<txid>', methods=['PUT'])
def api_update_transaksi(txid):
    payload = request.get_json(silent=True) or {}
    updates = payload.get('updates') or {}
    ok, message = update_transaction_fields(txid, updates, payload.get('user_edit') or 'Admin', payload.get('keterangan_edit') or 'Edit manual dari aplikasi')
    status = 200 if ok else 404
    return jsonify({'ok': ok, 'message': message}), status


@app.route('/api/nota-ocr', methods=['POST'])
def api_nota_ocr():
    try:
        payload = form_or_json()
        nota_name, nota_path = save_upload('foto_nota_ocr', 'nota_ocr')
        if not nota_name and not payload.get('nama_barang_ocr'):
            return jsonify({'ok': False, 'message': 'Upload nota atau isi data OCR manual terlebih dahulu.'}), 400

        qty = to_float(payload.get('qty_ocr'))
        harga = to_float(payload.get('harga_satuan_ocr'))
        total = to_float(payload.get('total_harga_ocr')) or qty * harga
        total_nota = to_float(payload.get('total_nota_ocr')) or total

        row = {
            'ID OCR': next_id('DATABASE_NOTA_OCR', 'OCR'),
            'Tanggal Upload': now_date(),
            'Jam Upload': now_time(),
            'Nama File Nota': nota_name,
            'Path Foto Nota': nota_path,
            'Tanggal Nota': payload.get('tanggal_nota') or '',
            'Nama Supplier / Toko': payload.get('supplier_ocr') or '',
            'Nomor Nota': payload.get('nomor_nota') or '',
            'Nama Barang OCR': payload.get('nama_barang_ocr') or '',
            'QTY OCR': qty,
            'Satuan OCR': payload.get('satuan_ocr') or '',
            'Harga Satuan OCR': harga,
            'Total Harga OCR': total,
            'Total Nota OCR': total_nota,
            'Status Pembacaan': 'MANUAL REVIEW',
            'Keterangan OCR': payload.get('keterangan_ocr') or 'Data OCR disiapkan untuk koreksi manual.'
        }
        append_dict_row('DATABASE_NOTA_OCR', row)
        return jsonify({'ok': True, 'message': f"Data nota/OCR tersimpan dengan ID {row['ID OCR']}", 'id': row['ID OCR']})
    except ValueError as exc:
        return jsonify({'ok': False, 'message': str(exc)}), 400
    except Exception as exc:
        return jsonify({'ok': False, 'message': f'Gagal menyimpan OCR: {exc}'}), 500


@app.route('/api/koreksi', methods=['POST'])
def api_koreksi():
    payload = request.get_json(silent=True) or {}
    row = {
        'ID Koreksi': next_id('LOG_KOREKSI_DATA', 'KOREKSI'),
        'Tanggal Koreksi': now_date(),
        'Jam Koreksi': now_time(),
        'ID Transaksi': payload.get('id_transaksi') or '',
        'Nama Barang Input': payload.get('nama_barang_input') or '',
        'Nama Barang Nota': payload.get('nama_barang_nota') or '',
        'QTY Input': to_float(payload.get('qty_input')),
        'QTY Nota': to_float(payload.get('qty_nota')),
        'Harga Satuan Input': to_float(payload.get('harga_satuan_input')),
        'Harga Satuan Nota': to_float(payload.get('harga_satuan_nota')),
        'Total Harga Input': to_float(payload.get('total_harga_input')),
        'Total Harga Nota': to_float(payload.get('total_harga_nota')),
        'Status Koreksi': payload.get('status_koreksi') or 'PERLU CEK',
        'Tindakan User': payload.get('tindakan_user') or '',
        'Keterangan': payload.get('keterangan') or ''
    }
    append_dict_row('LOG_KOREKSI_DATA', row)
    return jsonify({'ok': True, 'message': f"Koreksi tersimpan dengan ID {row['ID Koreksi']}", 'id': row['ID Koreksi']})


@app.route('/api/database-info')
def api_database_info():
    ensure_database()
    wb = load_workbook(DB_PATH, read_only=True)
    return jsonify({
        'path': str(DB_PATH).replace(str(BASE_DIR), ''),
        'exists': DB_PATH.exists(),
        'sheet_count': len(wb.sheetnames),
        'sheets': wb.sheetnames,
        'file_name': DB_PATH.name,
        'file_size_kb': round(DB_PATH.stat().st_size / 1024, 2)
    })


@app.route('/api/sheet/<sheet_name>')
def api_sheet(sheet_name):
    if sheet_name not in SHEET_HEADERS:
        return jsonify({'ok': False, 'message': 'Sheet tidak dikenal.'}), 404
    return jsonify({'ok': True, 'sheet': sheet_name, 'data': read_sheet(sheet_name)})


@app.route('/api/export-excel')
def api_export_excel():
    ensure_database()
    rebuild_stock()
    return send_file(DB_PATH, as_attachment=True, download_name='logistic_stock_database.xlsx')


@app.errorhandler(413)
def file_too_large(_):
    return jsonify({'ok': False, 'message': 'Ukuran upload terlalu besar. Maksimal total 32 MB per submit.'}), 413


if __name__ == '__main__':
    ensure_database()
    app.run(host='127.0.0.1', port=5000, debug=True)
